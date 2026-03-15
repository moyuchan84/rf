import pandas as pd
import numpy as np
import os
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# 설정
mockups_dir = "MockupFiles"
if not os.path.exists(mockups_dir):
    os.makedirs(mockups_dir)

def get_random_offset():
    return random.randint(1, 5), random.randint(1, 5) # row, col

def apply_basic_style(cell, is_header=False):
    if is_header:
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def create_complex_mockup(filename):
    wb = Workbook()
    
    # 1. History Sheet (항상 1번째)
    ws_hist = wb.active
    ws_hist.title = "History"
    r_off, c_off = get_random_offset()
    headers = ["Revision", "Date", "Author", "Comment"]
    for c_idx, h in enumerate(headers):
        cell = ws_hist.cell(row=r_off, column=c_off + c_idx, value=h)
        apply_basic_style(cell, True)
    
    for i in range(1, 6):
        row_data = [f"v1.{i}", f"2026-03-{10+i}", "User_A", f"Fixed issue {i}"]
        for c_idx, val in enumerate(row_data):
            cell = ws_hist.cell(row=r_off + i, column=c_off + c_idx, value=val)
            apply_basic_style(cell)

    # 2. Data Sheets (2번째부터)
    num_data_sheets = random.randint(1, 3)
    for s_idx in range(num_data_sheets):
        ws = wb.create_sheet(f"Data_Sheet_{s_idx+1}")
        r_off, c_off = get_random_offset()
        
        has_meta = random.choice([True, False])
        num_cols = random.randint(6, 14)
        num_rows = random.randint(20, 40)
        
        current_row = r_off
        
        # Meta 정보 섹션 (4-5줄)
        if has_meta:
            meta_keys = ["Process", "Step", "Machine", "Recipe", "LotID"]
            for i, key in enumerate(meta_keys):
                ws.cell(row=current_row + i, column=c_off, value=f"{key}:")
                ws.cell(row=current_row + i, column=c_off + 1, value=f"VAL_{random.randint(100,999)}")
            current_row += len(meta_keys) + 1 # 한 줄 띄움
        
        # Table Header
        # 비정형 컬럼명 생성 (중복, 누락, 세미콜론 포함)
        raw_headers = [f"Col_{i}" for i in range(num_cols)]
        raw_headers[0] = "Target_Name;Raw" # 세미콜론 포함
        raw_headers[1] = "Target_Name;Raw" # 중복 발생
        raw_headers[2] = "" # 누락 발생
        if num_cols > 5:
            raw_headers[5] = "Value;Offset"
            
        for c_idx, h in enumerate(raw_headers):
            cell = ws.cell(row=current_row, column=c_off + c_idx, value=h)
            apply_basic_style(cell, True)
            
        # Table Body
        for r_idx in range(1, num_rows + 1):
            for c_idx in range(num_cols):
                val = random.uniform(0, 1) if c_idx > 0 else f"K_{r_idx}"
                cell = ws.cell(row=current_row + r_idx, column=c_off + c_idx, value=val)
                apply_basic_style(cell)
                
    wb.save(os.path.join(mockups_dir, filename))

if __name__ == "__main__":
    for i in range(1, 6):
        create_complex_mockup(f"PHOTO_KEY_COMPLEX_V{i}.xlsx")
    print("5 Complex Mockup files created with Meta/Table structures.")
